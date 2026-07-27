import openpyxl
import sys

# Set standard output to UTF-8
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('../IN_Schemes_Master_Data_Collection_Template.xlsx')
print("Sheets in workbook:")
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"\nSheet: {sheet}")
    headers = [str(cell.value) if cell.value is not None else "" for cell in ws[1]]
    print(f"Headers: {headers}")
    first_row = [str(cell.value) if cell.value is not None else "" for cell in ws[2]]
    print(f"Row 2: {first_row}")
